import os
import re
import numpy as np
import pandas as pd


pd.set_option('expand_frame_repr',False)

ResultPathlist = [
r'F:\TestCase\ReportALLVersion\Result-4.0.14,Wed Jan  8 140819 2020',
    r'F:\TestCase\ReportALLVersion\Result-4.0.15,Thu Jan  9 143436 2020',
    r'F:\TestCase\ReportALLVersion\Result-4.0.16,Wed Jan 15 110452 2020'
]
summaryfilename = r'TestResult_summary.csv'
versionfilename = r'version.txt'
ResultALLfilename = r'F:\TestCase\ReportALLVersion\ResultALL.xlsx'

ResultALL = pd.DataFrame()
for ResultPath in ResultPathlist:
    summaryfilepath = os.path.join(ResultPath,summaryfilename)
    df_summary = pd.read_csv(summaryfilepath,encoding='utf_8_sig')
    with open(os.path.join(ResultPath,versionfilename),encoding='utf_8_sig') as f:
        version = ','.join(f.readline().split(',')[0:2])
        testtime = re.split(r'[\t,]',f.readline())[1]
    df_summary.insert(4,'version',version)
    df_summary.insert(5,'Test Time',testtime)
    df_summary.insert(0, 'Situation Number', list(range(len(df_summary))))
    #print(df_summary)
    ResultALL = pd.concat([ResultALL,df_summary]).reset_index(drop=True)

ResultALL = ResultALL.sort_values([ 'Situation Number','道路','光照','天气','车道线']).drop('Situation Number',axis=1).reset_index(drop=True)
print(ResultALL)

#ResultALL[['道路','光照','天气','车道线']][ResultALL[['道路','光照','天气','车道线']].duplicated()]=np.nan
sistuationnum = len(ResultALL[['道路','光照','天气','车道线']].drop_duplicates())
dup = ResultALL[['道路','光照','天气','车道线']].duplicated()
df_title = ResultALL[['道路','光照','天气','车道线']]
df_title[dup]=np.nan
ResultALL[['道路','光照','天气','车道线']] = df_title
print(ResultALL)
ResultALL.to_csv('ResultALL.csv',encoding='utf_8_sig',index=False)

from openpyxl import Workbook

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill,Border,Side,Alignment,Protection,Font

wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(ResultALL,index=False,header=True):
    ws.append(r)







border = Border(left=Side(border_style='thin',
                         color='FF000000'),
               right=Side(border_style='thin',
                          color='FF000000'),
              top=Side(border_style='thin',
                        color='FF000000'),
               bottom=Side(border_style='thin',
                           color='FF000000'),
              diagonal=Side(border_style='thin',
                             color='FF000000'),
               diagonal_direction=0,
               outline=Side(border_style='thin',
                          color='FF000000'),
               vertical=Side(border_style='thin',
                            color='FF000000'),
               horizontal=Side(border_style='thin',
                               color='FF000000')
               )

for row in ws.iter_cols():
    for cell in row:
        cell.border = border



font = Font(name='Calibri',
                 size=11,
               bold=True,
                italic=False,
               vertAlign=None,
              underline='none',
               strike=False,
               color='FF000000')

for row in ws.iter_cols( max_row=1):
    for cell in row:
        cell.font = font


fill = PatternFill(fill_type='solid',
               start_color='DCE6F1',
               end_color='DCE6F1')

for item in ws['A:D']:
    for cell in item:
        #print(cell)
        cell.fill = fill

for cell in  ws[1:1]:
    #print(cell)
    cell.fill = fill

resultnum = len(ResultPathlist)
for k in range(sistuationnum):
    lastresult = 'E'+str((k+1)*resultnum+1)+':'+'M'+str((k+1)*resultnum+1)
    for item in ws[lastresult]:
        for cell in item:
            #print(cell)
            fill = PatternFill(fill_type='solid',
                               start_color='FDE9D9',
                               end_color='FDE9D9')
            cell.fill = fill

columnwidthlist = []
for col in ws.iter_cols(values_only=True):
    width = 0
    for cell in col:
        #print(cell)
        try:
            valuelenth = len(cell.encode())+2
            if valuelenth>width:
                #print(len(cell))
                width = valuelenth
        except:
            if 6>width:
                width = 6
            pass
    columnwidthlist.append(width)



print(columnwidthlist)
colnames ='ABCDEFGHIJKLMNOPQRST'
for k in range(len(columnwidthlist)):
    width = columnwidthlist[k]
    ws.column_dimensions[colnames[k]].width = width


wb.save(ResultALLfilename)


